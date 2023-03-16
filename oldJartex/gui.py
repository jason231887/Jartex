from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('JartexStore')
root.geometry("1000x1000")

wb = Workbook()

wb = load_workbook('jartex.xlsx')

ws = wb.active

names = ws['A']
purchases = ws['B']

#Get user input for username to check
input_name = Entry(root)
input_name.pack(pady=20)
input_name.focus_set()

#Get the purchases made by the input username
def get_purchases():
    global input_name
    name = input_name.get()
    #Set name to lowercase to match in the list
    name.lower()
    list = ''
    #look through both columns and turn to tuple
    for cella, cellb in zip(names, purchases):
        #Set the names to lowercase to match the input username
        match = str(cella.value)
        match = match.lower()
        #If match, then add the purchase to list
        if match == name:
            list = f'{list + str(cellb.value)}\n'
    #Set label to the list of purchases
    label_name.config(text = list)

name_button = Button(root, text="Get Purchases", command= get_purchases)
name_button.pack(pady=20)

label_name = Label(root, text="")
label_name.pack(pady=20)

root.mainloop()